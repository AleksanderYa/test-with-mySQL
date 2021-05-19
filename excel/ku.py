import openpyxl
from datetime import date, datetime 


# Загружаем рабочий ексельфайл
wb = openpyxl.load_workbook(filename = 'Pricess2.xlsx')
# Загружаем рабочий лист
sheet_price1 = wb['Price1']
sheet_price2 = wb['Price2']
sheet_kurs = wb['Kurs']



# Полезно 
# sheet_kurs.max_row    -  максимум рядков, номер 
# sheet_kurs.min_row    -  максимум рядков, номер 
# sheet_kurs.max_column  максимум колонок, номер 
# sheet_kurs.min_column  максимум колонок, номер




#-----------------------------------------------------------------------------------------------------
def today():
    # Запись в дату текушую дату в формате 2020-11-08 
    data = str(datetime.today())
    today = data[:10].split('-')
    today_in = '{2},{1},{0}'.format(today[0][2:], today[1], today[2])
    today_org = '{2}-{1}-{0}'.format(today[0][2:], today[1], today[2])
    print(today_in)
    return [today_in, today_org]

#------Функция читает содержимое ячейки из листа "sheet_kurs" по существующим  рядкам--------------- 

def read_kurs(sheet=sheet_kurs):
    """Функция читает содержимое ячейки из листа "Kurs" по существующим  рядкам
    """
    # Итерируемся из списка мин до макс кол рядков  
    print('Start read kurs')
    for i in range(sheet_kurs.min_row, sheet_kurs.max_row + 1):
        # значение ячейки
        value = sheet.cell(i , 2).value
        value = str(value)
        print('read_kurs', value)
        # Если ячейка не пустая то заглянем туда
        if value != 'None' and value == today()[0]:
#             print('read_kurs_2', value)
            # Вызываем впомогательную фукцию которая проитерируется по колонкам и найдет значения
            res = read_col_kurs(i)
#             print(res)
            return res
#-----------------------------------------------------------------------------------------------------

#------Вспомагательная функция читает содержимое ячейки из "sheet_kurs" по существующим  колонкам------           

def read_col_kurs(row, sheet=sheet_kurs):
    # Список со значениями, куда мы будем записывать курсы
    read = []
    """Функция читает содержимое ячейки из листа "sheet_kurs" по существующим  колонкам    
       row - ряд в таблице
       sheet=sheet_kurs - лист екселя в котором делается работа.
    """
    # Итерируемся по непустым колонкам
    for ii in range(sheet_kurs.min_column, sheet_kurs.max_column):
    
        # Читаем значения ячейки начиная со второй, ибо первая это сама дата
        value = sheet.cell(row, ii+1).value
        # Добавляем значение в наш список
        read.append(value)
#         print('read_col_kurs', value)
    return read
    
    
read_kurses = read_kurs()
#-----------------------------------------------------------------------------------------------------

#------Функция читает содержимое из листа "sheet_price2" и определяет єто evro или dolars------------- 

def read_val(sheet=sheet_price2):
    """Функция читает содержимое ячейки из листа "Price2" по существующим  рядкам
    """
    print(sheet_price2.min_column, sheet_price2.max_column + 1)
    # Итерируемся из списка мин до макс кол рядков  
    for i in range(sheet_price2.min_column, sheet_price2.max_column + 1):
        # значение ячейки
        value = sheet.cell(i , 11).value
        if value == 'dolar':
            price = sheet.cell(i , 10).value
#             print(price*read_kurses[0])
            dolar(i, price)
#             print('yes, dolar')
        elif value == 'evro':
            price = sheet.cell(i , 10).value
            evro(i, price)
#             print('yes, evro')
            
#-----------------------------------------------------------------------------------------------------

def read_no_p1(sheet=sheet_price1):
    """
    """
    # Итерируемся из списка мин до макс кол рядков  
    for i in range(9 , sheet.max_row + 1):
        # значение ячейки
        value = sheet.cell(i , 19).value
#         print('read_no_p1-value', value)
#         print('read_no_p2-no_value', no_value)
        read_no_p2(value, i)
           
            
            
#-----------------------------------------------------------------------------------------------------

def read_no_p2(no, row, sheet=sheet_price1, sheet2=sheet_price2):
    """
    """
    # Итерируемся из списка мин до макс кол рядков  
    for i in range(sheet2.min_row, sheet2.max_row + 1):
        # значение ячейки
        no_valuee = sheet2.cell(i , 16).value
#         print('read_no_p2_novalue', no_valuee)
        valuee = sheet2.cell(i , 10).value
        if no == no_valuee:
            try:
                if str(sheet2.cell(i, 12).value) == '0':
                    price = sheet2.cell(i, 12).value
                    sheet.cell(row , 16).value = price
                else:
                    price = round((float(sheet2.cell(i, 12).value) + 0.2), 1)
                    sheet.cell(row , 16).value = price
            except:
                pass

#-----------------------------------------------------------------------------------------------------

def nds(val):
    val = (val/1.2)+0.005
    val =  round(val, 2)
    val = val*1.2
    res = float("%.3f" % (val))
    return res
    
#-----------------------------------------------------------------------------------------------------
#-----------------------------------------------------------------------------------------------------

# def input_kurs(
#     dolar=input(), 
#     evro=input()
# ):
#     dolar = float(dolar)
#     dolar_ots = dolar + 0.4
#     evro = float(evro)
#     evro_ots = evro + 0.4
#     write_kurs(evro, dolar, evro_ots, dolar_ots)
#     print('End input_kurs')

# def write_kurs(
#     evro, 
#     dolar,
#     evro_ots,
#     dolar_ots,
#     sheet=sheet_kurs
# ):
#     for i in range(sheet.min_row, sheet.max_row + 1):
#         cell = sheet.cell(i, 2).value
#         print('Start if')
#         if str(cell) == 'None':
#             sheet.cell(i, 2).value = today()[0]
#             sheet.cell(i, 3).value = dolar
#             sheet.cell(i, 4).value = dolar_ots
#             sheet.cell(i, 5).value = evro
#             sheet.cell(i, 6).value = evro_ots
#             save()
#-----------------------------------------------------------------------------------------------------

#-----------------------------------------------------------------------------------------------------

def evro(
    row, 
    price, 
    column_pre=12, 
    column_post=13,
    ws=sheet_price2
):
    try:
        ws.cell(row , column_pre).value = price*read_kurses[2]
        ws.cell(row , column_post).value = price*read_kurses[3]
#         save('Pricess2.xlsx')
        print(price,'ok')
    except Exception as e:
        print(e)    
        

#-----------------------------------------------------------------------------------------------------

def dolar(
    row,
    price, 
    column_pre=12, 
    column_post=13, 
    ws=sheet_price2
):

    try:
#         wbb = openpyxl.load_workbook(filename=path)
#         ws = wbb.worksheets[1]
#         value = 
        ws.cell(row , column_pre).value = price*read_kurses[0]
        ws.cell(row , column_post).value = price*read_kurses[1]
#         save('Pricess2.xlsx')
        print(price,'ok')
    except Exception as e:
        print(e)    
#-----------------------------------------------------------------------------------------------------

def write_info(
    row ,
    col,
    value, 
    ws=sheet_price2
):
    try:
        ws.cell(row , col).value = value
#         save('Pricess2.xlsx')
        print(ws.cell(row , col).value,'ok')
    except Exception as e:
        print(e)    
                
# def write(ryad , col, value):  
#         wb = openpyxl.load_workbook(filename='C:\\Users\\Sales2\Desktop\Pricess.xlsx' )
#         ws = wb.worksheets[0]
#         ws.cell(ryad , col).value = value
#         wb.save('C:\\Users\\Sales2\Desktop\Pricess.xlsx')
#         return print(ws.cell(ryad , col).value,'ok')
#-----------------------------------------------------------------------------------------------------
def save(name='Pricess2.xlsx'):
    try:
        wb.save(name)
        print("{0} is SAVE".format(name))
    except Exception as e:
        print(e)

#-----------------------------------------------------------------------------------------------------
def write_all_info():    
    try:
        write_info(21, 4, today()[0])
        write_info(22, 6, read_kurses[0])
        write_info(22, 7, read_kurses[1])
        write_info(23, 6, read_kurses[2])
        write_info(23, 7, read_kurses[3])
        write_info(32, 11, today()[1] ,ws=sheet_price1)
    except Exception as e:
        print(e)
#-----------------------------------------------------------------------------------------------------    
def main():
    read_val()
    write_all_info()
    read_no_p1()
    save()
    
if __name__ == '__main__':
    main()
